import pandas as pd
import numpy as np
from openpyxl import load_workbook
from os import getcwd
import re
import time
import torch
from torch import nn
import transformers as ppb
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score, precision_recall_fscore_support, roc_auc_score, multilabel_confusion_matrix
import matplotlib.pyplot as plt


EPOCH = 80 
BATCH_SIZE = 32 
MAX_LENGTH = 512


class BertClassifier(nn.Module):
    def __init__(self, num_labels):
        super().__init__()
        self.bert = ppb.BertModel.from_pretrained(pretrained_weights)
        self.classifier = nn.Linear(self.bert.config.hidden_size, num_labels)

    def forward(self, input_ids, attention_mask=None):
        outputs = self.bert(input_ids, attention_mask=attention_mask)
        logits = self.classifier(outputs.last_hidden_state[:, 0, :])
        return logits




# function to predict the label on a given text in excel workbook
def process_excel_file(input_file, output_file, model, tokenizer, batch_size=BATCH_SIZE, max_length=MAX_LENGTH):
    wb = load_workbook(input_file)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Find the "text" column
        text_col_idx = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "text":
                text_col_idx = col
                break

        if text_col_idx is None:
            print(f"No 'text' column found in sheet '{sheet_name}'. Skipping this sheet.")
            continue

        # Pre-process and predict labels for the "text" column in batches
        texts = [ws.cell(row=row, column=text_col_idx).value for row in range(2, ws.max_row + 1)]
        processed_texts = [re.sub(r'[^\u0000-\uFFFF]+', ' ', str(text)) for text in texts]

        num_batches = int(np.ceil(len(processed_texts) / batch_size))
        predictions = []
        for i in range(num_batches):
            batch_texts = processed_texts[i * batch_size:(i + 1) * batch_size]

            encoded_input = tokenizer.batch_encode_plus(
                batch_texts,
                add_special_tokens=True,
                max_length=max_length,
                padding='max_length',
                truncation=True,
                return_attention_mask=True,
                return_tensors='pt'
            )

            input_ids = encoded_input['input_ids'].to(device)
            attention_mask = encoded_input['attention_mask'].to(device)

            with torch.no_grad():
                logits = model(input_ids, attention_mask=attention_mask)
                batch_predictions = torch.argmax(logits, dim=1).cpu().numpy()
                predictions.extend(batch_predictions)

        # Create a new "Prediction" column and save the predictions
        pred_col_idx = ws.max_column + 1
        ws.cell(row=1, column=pred_col_idx).value = "Prediction"
        for row, prediction in enumerate(predictions, start=2):
            ws.cell(row=row, column=pred_col_idx).value = label_decoder[prediction]

    # Save the workbook to a new file
    wb.save(output_file)


# Load the Tokenizer
tokenizer_class = ppb.BertTokenizer
pretrained_weights = 'bert-base-uncased'
tokenizer = tokenizer_class.from_pretrained(pretrained_weights)

torch.manual_seed(3407)

# Load the dataset
df = pd.read_csv('xxxxxxxxxxxx.csv')

# Filter the text column to keep only normal characters
df['text'] = df['text'].str.replace(r'[^\u0000-\uFFFF]+', ' ', regex=True)

# Create a label encoder for the concepts, including a special label for irrelevant texts
concepts = ['fair value', 'deal price', 'discounted cash flow', 'market efficiency',
            'stock price', 'market multiples', 'analyst price target', 'multiples-based valuation', 
            'abnormal return']

label_encoder = {c: i for i, c in enumerate(concepts)}
label_encoder['irrelevant'] = len(concepts)

# Encode the labels
df['label'] = df['concept'].map(label_encoder).fillna(label_encoder['irrelevant']).astype(int)

# Decoder
label_decoder = {i: c for c, i in label_encoder.items()}

# Initialize the custom BERT classifier
model = BertClassifier(len(label_encoder))

# Set up pyTorch
device = torch.device('cuda:1' if torch.cuda.is_available() else 'cpu')
print(f"Current Device: {device} \n")
model.to(device)


# Prepare the input data for the BERT model
batch_size = BATCH_SIZE
max_length = MAX_LENGTH

encoded_input = tokenizer.batch_encode_plus(
    df['text'].values,
    add_special_tokens=True,
    max_length=max_length,
    padding='max_length',
    truncation=True,
    return_attention_mask=True,
    return_tensors='pt'
)

input_ids = encoded_input['input_ids']
attention_mask = encoded_input['attention_mask']

num_examples = len(df)
remainder = num_examples % batch_size
if remainder != 0:
    num_pad_examples = batch_size - remainder
    input_ids = np.pad(input_ids, ((0, num_pad_examples), (0, 0)), 'constant', constant_values=tokenizer.pad_token_id)
    attention_mask = np.pad(attention_mask, ((0, num_pad_examples), (0, 0)), 'constant', constant_values=0) 

padding_df = pd.DataFrame({'label': [-1] * num_pad_examples})
df = pd.concat([df, padding_df], ignore_index=True)

# Split the data into train and test sets
X_train, X_test, y_train, y_test = train_test_split(
    input_ids, df['label'], 
    test_size=0.2, 
    stratify=df['label'], 
    random_state=42
)


# Define the training loop
epochs = EPOCH
optimizer = torch.optim.Adam(model.parameters(), lr=1e-5)

loss_fn = torch.nn.CrossEntropyLoss()

train_losses = []
train_accs = []
test_losses = []
test_accs = []
test_f1s= []

# Initialize an empty DataFrame to store the metrics
num_labels = len(label_encoder)
columns = ['Epoch', 'Accuracy', 'Precision', 'Recall', 'F1 score', 'AUC score', 'Avg False Positive', 'Avg False Negative']
columns += [f'{label}_TP' for label in label_encoder] + [f'{label}_TN' for label in label_encoder] + [f'{label}_FP' for label in label_encoder] + [f'{label}_FN' for label in label_encoder]

metrics_df = pd.DataFrame(columns=columns)

# Run epochs 
for epoch in range(epochs):
    start_time = time.time() 

    # Train the model
    model.train()
    train_loss = 0
    train_acc = 0
    num_batches = len(X_train) // batch_size
    for i in range(num_batches):
        start_index = i * batch_size
        end_index = start_index + batch_size
        
        # Filter out padded examples
        valid_indices = np.where(y_train.values[start_index:end_index] != -1)[0]
        
        batch_input_ids = torch.tensor(X_train[start_index:end_index][valid_indices]).to(device)
        batch_attention_mask = torch.tensor(attention_mask[start_index:end_index][valid_indices]).to(device)
        batch_labels = torch.LongTensor(y_train.values[start_index:end_index][valid_indices]).to(device)

        optimizer.zero_grad()

        logits = model(batch_input_ids, attention_mask=batch_attention_mask)
        loss = loss_fn(logits, batch_labels)
        loss.backward()
        optimizer.step()

        train_loss += loss.item()
        train_acc += accuracy_score(batch_labels.cpu(), torch.argmax(logits, dim=1).cpu())

    train_loss /= num_batches
    train_acc /= num_batches
    train_losses.append(train_loss)
    train_accs.append(train_acc)

    # Evaluate the model on the test set
    model.eval()
    test_loss = 0
    test_acc = 0
    y_true = []
    y_pred = []
    num_batches = len(X_test) // batch_size
    with torch.no_grad():
        for i in range(num_batches):
            start_index = i * batch_size
            end_index = start_index + batch_size
            
            # Filter out padded examples
            valid_indices = np.where(y_test.values[start_index:end_index] != -1)[0]
            
            batch_input_ids = torch.tensor(X_test[start_index:end_index][valid_indices]).to(device)
            batch_attention_mask = torch.tensor(attention_mask[start_index:end_index][valid_indices]).to(device)
            batch_labels = torch.LongTensor(y_test.values[start_index:end_index][valid_indices]).to(device)

            logits = model(batch_input_ids, attention_mask=batch_attention_mask)
            loss = loss_fn(logits, batch_labels)

            test_loss += loss.item()
            batch_logits = torch.argmax(logits, dim=1)
            test_acc += accuracy_score(batch_labels.cpu(), batch_logits.cpu())
            
            y_true.extend(batch_labels.cpu().numpy())
            y_pred.extend(batch_logits.cpu().numpy())

    test_loss /= num_batches
    test_acc /= num_batches
    test_losses.append(test_loss)
    test_accs.append(test_acc)

    # Calculate evaluation metrics
    precision, recall, f1, _ = precision_recall_fscore_support(y_true, y_pred, average='weighted', zero_division=1)

    all_labels = list(label_encoder.keys())


    #Debugging
    print("Unique y_true:", np.unique(np.array(y_true)))
    print("Unique y_pred:", np.unique(np.array(y_pred)))
    print("All possible labels:", all_labels)
    
    auc = roc_auc_score(pd.get_dummies(np.array(y_true), columns=all_labels), pd.get_dummies(np.array(y_pred), columns=all_labels), multi_class='ovr')
    
    # Calculate average False Positive and False Negative rates
    mcm = multilabel_confusion_matrix(y_true, y_pred, labels=range(num_labels))
    avg_fp = np.mean([matrix[0, 1] for matrix in mcm])
    avg_fn = np.mean([matrix[1, 0] for matrix in mcm])
    
    # Extract TP, TN, FP, FN for each label
    tp_values = [matrix[1, 1] for matrix in mcm]
    tn_values = [matrix[0, 0] for matrix in mcm]
    fp_values = [matrix[0, 1] for matrix in mcm]
    fn_values = [matrix[1, 0] for matrix in mcm]

    # Save evaluation metrics to DataFrame
    metric_values = [epoch + 1, test_acc, precision, recall, f1, auc if auc is not None else 'N/A', avg_fp, avg_fn] + tp_values + tn_values + fp_values + fn_values
    temp_df = pd.DataFrame([metric_values], columns=columns)
    metrics_df = pd.concat([metrics_df, temp_df], ignore_index=True)
    metrics_df.to_csv(getcwd() + '/evaluation_metrics_80.csv', index=False)

    runtime = time.time() - start_time
    print(f'Epoch {epoch+1}, train loss: {train_loss:.4f}, train acc: {train_acc:.4f}, test loss: {test_loss:.4f}, test acc: {test_acc:.4f}\nTime elapsed: {(runtime // 60):.0f} minutes, {(runtime % 60):.0f} seconds\n')


metrics_df.to_csv(getcwd() + '/evaluation_metrics_80.csv', index=False)


# Plot the training and test losses, accuracies, and F1 scores
fig, axs = plt.subplots(1, 3, figsize=(15, 5))
axs[0].plot(train_losses, label='train')
axs[0].plot(test_losses, label='test')
axs[0].legend()
axs[0].set_title('Loss')
axs[0].set_xlabel('Epochs')
# axs[0].set_ylabel('Loss')
# for i, loss in enumerate(train_losses):
#     axs[0].text(i, loss, f'{loss:.2f}')
# for i, loss in enumerate(test_losses):
#     axs[0].text(i, loss, f'{loss:.2f}')

axs[1].plot(train_accs, label='train')
axs[1].plot(test_accs, label='test')
axs[1].legend()
axs[1].set_title('Accuracy')
axs[1].set_xlabel('Epochs')
# axs[1].set_ylabel('Accuracy')
# for i, acc in enumerate(train_accs):
#     axs[1].text(i, acc, f'{acc:.2f}')
# for i, acc in enumerate(test_accs):
#     axs[1].text(i, acc, f'{acc:.2f}')

axs[2].plot(test_f1s, label='test')
axs[2].legend()
axs[2].set_title('F1 Score for Test Set')
axs[2].set_xlabel('Epochs')

fig.suptitle('Progression for BERT Epochs Implemented by Torch')

plt.savefig(getcwd() + '/training_and_test_progression_80.png', dpi=800)

# plt.show()


# Example usage
input_excel_file = getcwd() + "/prediction/test_reports.xlsx"
output_excel_file = getcwd() + "/prediction/test_reports_prediction.xlsx"

process_excel_file(input_excel_file, output_excel_file, model, tokenizer)